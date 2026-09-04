from io import BytesIO

import pandas as pd

from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)

from openpyxl.utils import get_column_letter


def create_excel_report(report_data):

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        for sheet_name, dataframe in report_data.items():

            # Excel sheet names cannot exceed 31 characters
            safe_sheet_name = sheet_name[:31]

            dataframe.to_excel(
                writer,
                sheet_name=safe_sheet_name,
                index=False
            )

            worksheet = writer.book[
                safe_sheet_name
            ]

            # =================================================
            # HEADER FORMATTING
            # =================================================

            header_fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7"
            )

            header_font = Font(
                bold=True
            )

            thin_border = Border(
                bottom=Side(
                    style="thin",
                    color="B7B7B7"
                )
            )

            for cell in worksheet[1]:

                cell.font = header_font
                cell.fill = header_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                cell.border = thin_border

            # =================================================
            # FREEZE HEADER
            # =================================================

            worksheet.freeze_panes = "A2"

            # =================================================
            # FILTER
            # =================================================

            if worksheet.max_column > 0:

                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

            # =================================================
            # AUTO WIDTH
            # =================================================

            for column_cells in worksheet.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:

                    try:

                        value = (
                            ""
                            if cell.value is None
                            else str(cell.value)
                        )

                        if len(value) > max_length:
                            max_length = len(value)

                    except Exception:
                        pass

                adjusted_width = min(
                    max(
                        max_length + 2,
                        12
                    ),
                    45
                )

                worksheet.column_dimensions[
                    column_letter
                ].width = adjusted_width

            # =================================================
            # SUMMARY SHEET SPECIAL FORMATTING
            # =================================================

            if safe_sheet_name == "Summary":

                worksheet.column_dimensions[
                    "A"
                ].width = 38

                worksheet.column_dimensions[
                    "B"
                ].width = 16

                for row in worksheet.iter_rows(
                    min_row=2
                ):

                    row[1].alignment = Alignment(
                        horizontal="center"
                    )

    output.seek(0)

    return output.getvalue()
