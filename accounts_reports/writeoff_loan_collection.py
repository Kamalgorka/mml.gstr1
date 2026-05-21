import os
import re
import tempfile
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


MONTH_ORDER = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
]


def norm_col(value):
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace(".", "")
    )


def clean_value(value):
    if value is None:
        return ""

    return str(value).strip()


def to_number(value):
    if value is None:
        return 0

    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if value in ["", "-", "nan", "NaN"]:
            return 0

    try:
        return float(value)
    except Exception:
        return 0


def display_value(value):
    value = round(float(value), 0)

    if value == 0:
        return "-"

    return int(value)


def next_month_name(month_name):
    idx = MONTH_ORDER.index(month_name)
    return MONTH_ORDER[(idx + 1) % 12]


def get_excel_engine(file):
    name = getattr(file, "name", str(file)).lower()

    if name.endswith(".xlsb"):
        return "pyxlsb"

    return None


def prepare_repayment_summary(repayment_file):
    engine = get_excel_engine(repayment_file)

    df = pd.read_excel(repayment_file, engine=engine)

    rename_map = {}

    for col in df.columns:
        n = norm_col(col)

        if n == "loanid":
            rename_map[col] = "loan_id"
        elif n == "principalcollected":
            rename_map[col] = "principal_collected"
        elif n == "interestcollected":
            rename_map[col] = "interest_collected"
        elif n in ["waiveoffamt", "waiveoffamount", "waiveoff"]:
            rename_map[col] = "waiveoff_amt"

    df = df.rename(columns=rename_map)

    required_cols = [
        "loan_id",
        "principal_collected",
        "interest_collected",
        "waiveoff_amt",
    ]

    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        raise Exception(f"Missing columns in repayment file: {missing}")

    df["loan_id"] = df["loan_id"].astype(str).str.strip()

    for col in ["principal_collected", "interest_collected", "waiveoff_amt"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    summary = (
        df.groupby("loan_id", as_index=False)
        .agg(
            principal_collected=("principal_collected", "sum"),
            interest_collected=("interest_collected", "sum"),
            waiveoff_amt=("waiveoff_amt", "sum"),
        )
    )

    return summary.set_index("loan_id").to_dict("index")


def find_header_row_and_loan_col(ws):
    for row in range(1, 15):
        for col in range(1, ws.max_column + 1):
            value = clean_value(ws.cell(row=row, column=col).value)

            if norm_col(value) == "loanid":
                return row, col

    raise Exception(f"Loan ID column not found in sheet: {ws.title}")


def find_previous_closing_col(ws, header_row):
    closing_cols = []

    for col in range(1, ws.max_column + 1):
        value = clean_value(ws.cell(row=header_row, column=col).value)

        match = re.match(r"Closing\s+([A-Za-z]{3,})", value, re.IGNORECASE)

        if match:
            month = match.group(1)[:3].title()

            if month in MONTH_ORDER:
                closing_cols.append((col, month))

    if not closing_cols:
        raise Exception(f"No previous closing month column found in sheet: {ws.title}")

    return closing_cols[-1]


def find_last_data_row(ws, loan_col, header_row):
    last_row = header_row
    blank_count = 0

    for row in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=loan_col).value

        if value is not None and str(value).strip() != "":
            last_row = row
            blank_count = 0
        else:
            blank_count += 1

        # Stop scanning after 200 continuous blank Loan ID rows
        if blank_count >= 200 and last_row > header_row:
            break

    return last_row


def update_sheet(ws, repayment_lookup):
    header_row, loan_col = find_header_row_and_loan_col(ws)

    previous_closing_col, previous_month = find_previous_closing_col(ws, header_row)

    current_month = next_month_name(previous_month)

    start_col = previous_closing_col + 1

    month_heading_row = header_row - 2
    total_row = header_row - 1

    headers = [
        f"P.A {current_month}",
        f"I.A {current_month}",
        f"W.A {current_month}",
        f"Closing {current_month}",
    ]

    # Remove old merged range if target area already merged
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row == month_heading_row
            and merged_range.min_col >= start_col
            and merged_range.min_col <= start_col + 3
        ):
            ws.unmerge_cells(str(merged_range))

    # Month heading
    ws.cell(month_heading_row, start_col).value = f"{current_month}-{datetime.now().strftime('%y')}"
    ws.merge_cells(
        start_row=month_heading_row,
        start_column=start_col,
        end_row=month_heading_row,
        end_column=start_col + 3,
    )

    ws.cell(month_heading_row, start_col).font = Font(bold=True)
    ws.cell(month_heading_row, start_col).alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Header row
    for i, header in enumerate(headers):
        cell = ws.cell(header_row, start_col + i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[cell.column_letter].width = 14

    last_data_row = find_last_data_row(ws, loan_col, header_row)

    total_pa = 0
    total_ia = 0
    total_wa = 0
    total_closing = 0

    last_data_row = find_last_data_row(ws, loan_col, header_row)
        loan_id = ws.cell(row=row, column=loan_col).value

        if loan_id is None or str(loan_id).strip() == "":
            continue

        loan_id = str(loan_id).strip()

        data = repayment_lookup.get(
            loan_id,
            {
                "principal_collected": 0,
                "interest_collected": 0,
                "waiveoff_amt": 0,
            }
        )

        pa = data["principal_collected"]
        ia = data["interest_collected"]
        wa = data["waiveoff_amt"]

        previous_closing = to_number(ws.cell(row=row, column=previous_closing_col).value)
        closing = previous_closing - pa

        values = [
            display_value(pa),
            display_value(ia),
            display_value(wa),
            display_value(closing),
        ]

        for i, value in enumerate(values):
            cell = ws.cell(row=row, column=start_col + i)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")

        total_pa += pa
        total_ia += ia
        total_wa += wa
        total_closing += closing

    totals = [
        display_value(total_pa),
        display_value(total_ia),
        display_value(total_wa),
        display_value(total_closing),
    ]

    for i, value in enumerate(totals):
        cell = ws.cell(row=total_row, column=start_col + i)
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    return current_month


def process_writeoff_loan_collection(
    writeoff_file,
    repayment_file,
    output_dir=None,
    progress_callback=None
):
    if progress_callback:
        progress_callback(10, "Reading repayment file...")

    repayment_lookup = prepare_repayment_summary(repayment_file)

    if progress_callback:
        progress_callback(30, "Opening WriteOff loan collection file...")

    wb = load_workbook(
        writeoff_file,
        data_only=False,
        keep_links=False
    )

    required_sheets = ["System Write-Off", "Manual Write-Off"]

    missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]

    if missing_sheets:
        raise Exception(f"Missing sheets in WriteOff file: {missing_sheets}")

    if progress_callback:
        progress_callback(45, "Updating System Write-Off sheet...")

    system_month = update_sheet(wb["System Write-Off"], repayment_lookup)

    if progress_callback:
        progress_callback(70, "Updating Manual Write-Off sheet...")

    manual_month = update_sheet(wb["Manual Write-Off"], repayment_lookup)

    if output_dir is None:
        output_dir = tempfile.gettempdir()

    output_file = os.path.join(
        output_dir,
        f"WriteOff_Loan_Collection_Updated_{system_month}.xlsx"
    )

    if progress_callback:
        progress_callback(90, "Saving updated workbook...")

    wb.save(output_file)

    if progress_callback:
        progress_callback(100, "Report generated successfully.")

    return output_file
