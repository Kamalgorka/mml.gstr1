import os
import zipfile
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def process_arc_writeoff_entries(uploaded_file, output_dir, progress_callback=None):

    if progress_callback:
        progress_callback(10, "Reading uploaded Excel file")

    output_folder = os.path.join(output_dir, "ARC_WriteOff_Entries")
    os.makedirs(output_folder, exist_ok=True)

    xls = pd.ExcelFile(uploaded_file)

    summary = []
    consolidated_list = []

    required_cols = [
        "B.Code",
        "P.C",
        "Principal Amount",
        "Interest Amount",
        "OTS Amount"
    ]

    total_sheets = len(xls.sheet_names)

    for idx, sheet_name in enumerate(xls.sheet_names, start=1):

        if progress_callback:
            progress_callback(
                int(10 + (idx / total_sheets) * 60),
                f"Processing sheet: {sheet_name}"
            )

        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
        df.columns = df.columns.astype(str).str.strip()

        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise Exception(
                f"Sheet '{sheet_name}' missing required columns: {missing_cols}"
            )

        df = df[required_cols].copy()

        df["B.Code"] = df["B.Code"].astype(str).str.strip()
        df["P.C"] = df["P.C"].astype(str).str.strip().str.upper()

        for col in ["Principal Amount", "Interest Amount", "OTS Amount"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        df = df[df["B.Code"].notna()]
        df = df[df["B.Code"] != ""]
        df = df[df["B.Code"].str.lower() != "nan"]
        df = df[df["P.C"].isin(["JLG", "CPP", "IL"])]

        output_df = build_arc_output_df(df)

        safe_sheet_name = clean_file_name(sheet_name)
        file_path = os.path.join(output_folder, f"{safe_sheet_name}_Entry.xlsx")

        write_arc_entry_format(output_df, file_path, sheet_name)

        consolidated_list.append(output_df)

        summary.append({
            "sheet_name": sheet_name,
            "rows": len(output_df)
        })

    if progress_callback:
        progress_callback(80, "Preparing consolidated file")

    consolidated_df = pd.concat(consolidated_list, ignore_index=True)

    consolidated_df = consolidated_df.groupby("B.Code", as_index=False).agg({
        "Principal_JLG": "sum",
        "Principal_CPP": "sum",
        "Principal_IL": "sum",
        "Interest_JLG": "sum",
        "Interest_CPP": "sum",
        "Interest_IL": "sum",
        "OTS Amount": "sum"
    })

    consolidated_path = os.path.join(
        output_folder,
        "Consolidated_ARC_WriteOff_Entry.xlsx"
    )

    write_arc_entry_format(
        consolidated_df,
        consolidated_path,
        "Consolidated"
    )

    summary.append({
        "sheet_name": "Consolidated",
        "rows": len(consolidated_df)
    })

    if progress_callback:
        progress_callback(90, "Creating ZIP file")

    zip_path = os.path.join(
        output_dir,
        "ARC_WriteOff_Entries_Output.zip"
    )

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for file_name in os.listdir(output_folder):
            file_path = os.path.join(output_folder, file_name)
            zipf.write(file_path, arcname=file_name)

    if progress_callback:
        progress_callback(100, "Completed")

    return summary, zip_path


def build_arc_output_df(df):

    grouped = df.groupby(["B.Code", "P.C"], as_index=False).agg({
        "Principal Amount": "sum",
        "Interest Amount": "sum",
        "OTS Amount": "sum"
    })

    output_rows = []

    for bcode in sorted(grouped["B.Code"].unique()):

        row = {
            "B.Code": bcode,
            "Principal_JLG": 0,
            "Principal_CPP": 0,
            "Principal_IL": 0,
            "Interest_JLG": 0,
            "Interest_CPP": 0,
            "Interest_IL": 0,
            "OTS Amount": 0
        }

        branch_df = grouped[grouped["B.Code"] == bcode]

        for _, r in branch_df.iterrows():
            pc = r["P.C"]

            row[f"Principal_{pc}"] += r["Principal Amount"]
            row[f"Interest_{pc}"] += r["Interest Amount"]
            row["OTS Amount"] += r["OTS Amount"]

        output_rows.append(row)

    return pd.DataFrame(output_rows)


def clean_file_name(name):
    invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    name = str(name).strip()

    for ch in invalid_chars:
        name = name.replace(ch, "_")

    return name[:80]


def write_arc_entry_format(df, file_path, sheet_title):

    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_title)[:31]

    blue_fill = PatternFill("solid", fgColor="BFEAF5")
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:D1")
    ws.merge_cells("E1:G1")
    ws.merge_cells("H1:H2")

    ws["A1"] = "B.Code"
    ws["B1"] = "Principal Amount"
    ws["E1"] = "Interest Amount"
    ws["H1"] = "OTS Amount"

    headers_row2 = {
        "B2": "JLG",
        "C2": "CPP",
        "D2": "IL",
        "E2": "JLG",
        "F2": "CPP",
        "G2": "IL"
    }

    for cell, value in headers_row2.items():
        ws[cell] = value

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.font = bold_font
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = border

    start_row = 3

    if df.empty:
        df = pd.DataFrame([{
            "B.Code": "",
            "Principal_JLG": 0,
            "Principal_CPP": 0,
            "Principal_IL": 0,
            "Interest_JLG": 0,
            "Interest_CPP": 0,
            "Interest_IL": 0,
            "OTS Amount": 0
        }])

    for idx, row in df.iterrows():

        excel_row = start_row + idx

        values = [
            row["B.Code"],
            row["Principal_JLG"],
            row["Principal_CPP"],
            row["Principal_IL"],
            row["Interest_JLG"],
            row["Interest_CPP"],
            row["Interest_IL"],
            row["OTS Amount"]
        ]

        for col_no, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_no)
            cell.value = value
            cell.alignment = center
            cell.border = border

    total_row = start_row + len(df)

    ws.cell(row=total_row, column=1).value = "Grand Total"

    for col in range(2, 9):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = (
            f"=SUM({col_letter}{start_row}:{col_letter}{total_row - 1})"
        )

    for col in range(1, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.font = bold_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions["A"].width = 14

    ws.freeze_panes = "A3"

    wb.save(file_path)
