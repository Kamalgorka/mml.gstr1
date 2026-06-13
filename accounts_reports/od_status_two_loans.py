import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HUB_MAP = {
    "BH1": "Hub-1 Chandigarh",
    "BH2": "Hub-2 Ahmedabad",
    "BH3": "Hub-3 Patna",
    "BH4": "Hub-4 Lucknow",
    "BH5": "Hub-5 Kolkata",
    "BH6": "Hub-6 Bhubaneswar",
}


def clean_col(col):
    return str(col).strip().lower().replace(" ", "_")


def process_od_status_two_loans(uploaded_file, output_dir, progress_callback=None):

    if progress_callback:
        progress_callback(10, "Reading input file...")

    df = pd.read_excel(uploaded_file)
    df.columns = [clean_col(c) for c in df.columns]

    required_cols = [
        "od_days",
        "column1",
        "cluster",
        "total_arrear",
        "outstanding_principal",
        "finpage_loan_no",
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    if progress_callback:
        progress_callback(30, "Preparing loan status...")

    df["od_days"] = pd.to_numeric(df["od_days"], errors="coerce").fillna(0)
    df["total_arrear"] = pd.to_numeric(df["total_arrear"], errors="coerce").fillna(0)
    df["outstanding_principal"] = pd.to_numeric(df["outstanding_principal"], errors="coerce").fillna(0)

    od_index = list(df.columns).index("od_days") + 1
    df.insert(
        od_index,
        "loan_status",
        df["od_days"].apply(lambda x: "Regular Loan" if x == 0 else "Arrear Loan")
    )

    df["hub_name"] = df["column1"].astype(str).str.strip().str.upper().map(HUB_MAP).fillna(df["column1"])
    df["cluster"] = df["cluster"].astype(str).str.strip()

    if progress_callback:
        progress_callback(55, "Preparing summary sheet...")

    summary_rows = []

    hub_order = list(HUB_MAP.values())

    for hub in hub_order:
        hub_df = df[df["hub_name"] == hub]
        if hub_df.empty:
            continue

        summary_rows.append({
            "HUB/ Region": hub,
            "Arrear Amount": "",
            "PAR Amount": "",
            "Arrear No. of Loans": "",
            "Regular Arrear Amount": "",
            "Regular PAR Amount": "",
            "Regular No. of Loans": "",
            "row_type": "hub"
        })

        for region in sorted(hub_df["cluster"].dropna().unique()):
            region_df = hub_df[hub_df["cluster"] == region]

            arrear_df = region_df[region_df["loan_status"] == "Arrear Loan"]
            regular_df = region_df[region_df["loan_status"] == "Regular Loan"]

            summary_rows.append({
                "HUB/ Region": region,
                "Arrear Amount": arrear_df["total_arrear"].sum(),
                "PAR Amount": arrear_df["outstanding_principal"].sum(),
                "Arrear No. of Loans": arrear_df["finpage_loan_no"].count(),
                "Regular Arrear Amount": "-",
                "Regular PAR Amount": "-",
                "Regular No. of Loans": regular_df["finpage_loan_no"].count(),
                "row_type": "region"
            })

        arrear_total = hub_df[hub_df["loan_status"] == "Arrear Loan"]
        regular_total = hub_df[hub_df["loan_status"] == "Regular Loan"]

        summary_rows.append({
            "HUB/ Region": f"{hub} Total",
            "Arrear Amount": arrear_total["total_arrear"].sum(),
            "PAR Amount": arrear_total["outstanding_principal"].sum(),
            "Arrear No. of Loans": arrear_total["finpage_loan_no"].count(),
            "Regular Arrear Amount": "-",
            "Regular PAR Amount": "-",
            "Regular No. of Loans": regular_total["finpage_loan_no"].count(),
            "row_type": "total"
        })

    summary_df = pd.DataFrame(summary_rows)

    grand_arrear = df[df["loan_status"] == "Arrear Loan"]
    grand_regular = df[df["loan_status"] == "Regular Loan"]

    grand_total = pd.DataFrame([{
        "HUB/ Region": "Grand Total",
        "Arrear Amount": grand_arrear["total_arrear"].sum(),
        "PAR Amount": grand_arrear["outstanding_principal"].sum(),
        "Arrear No. of Loans": grand_arrear["finpage_loan_no"].count(),
        "Regular Arrear Amount": "-",
        "Regular PAR Amount": "-",
        "Regular No. of Loans": grand_regular["finpage_loan_no"].count(),
        "row_type": "grand"
    }])

    summary_df = pd.concat([summary_df, grand_total], ignore_index=True)
    # Create output folder if not exists
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "OD_Status_of_Two_Loans_Report.xlsx")

    if progress_callback:
        progress_callback(75, "Writing Excel output...")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Final Sheet")
        summary_df.drop(columns=["row_type"]).to_excel(
            writer, index=False, sheet_name="Summary", startrow=2
        )

    format_summary(output_file, summary_df)

    if progress_callback:
        progress_callback(100, "Report generated successfully.")

    return output_file


def format_summary(output_file, summary_df):
    wb = load_workbook(output_file)
    ws = wb["Summary"]

    ws.merge_cells("A1:G1")
    ws["A1"] = "Member with OD & Regular Loans Summary"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:D2")
    ws["B2"] = "Arrear Loan"

    ws.merge_cells("E2:G2")
    ws["E2"] = "Regular Loan"

    ws["A2"] = "HUB/ Region"
    ws["A3"] = "HUB/ Region"

    headers = [
        "Arrear Amount", "PAR Amount", "No. of Loans",
        "Arrear Amount", "PAR Amount", "No. of Loans"
    ]

    for col, value in enumerate(headers, start=2):
        ws.cell(row=3, column=col).value = value

    ws.delete_rows(4)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="000000")
    blue_side = Side(style="thin", color="4F81BD")

    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for row in [1, 2, 3]:
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    color_map = {
        "Hub-1 Chandigarh": "D8E4BC",
        "Hub-2 Ahmedabad": "F2DCDB",
        "Hub-3 Patna": "DCEAF7",
        "Hub-4 Lucknow": "E4DFEC",
        "Hub-5 Kolkata": "FFFFCC",
        "Hub-6 Bhubaneswar": "D9D9D9",
    }

    start_row = 4
    for idx, row_data in summary_df.iterrows():
        excel_row = start_row + idx
        row_type = row_data["row_type"]
        hub_region = row_data["HUB/ Region"]

        fill_color = "FFFFFF"

        for hub_name, color in color_map.items():
            if hub_name in str(hub_region):
                fill_color = color
                break

        for col in range(1, 8):
            ws.cell(excel_row, col).fill = PatternFill("solid", fgColor=fill_color)

        if row_type in ["hub", "total", "grand"]:
            for col in range(1, 8):
                ws.cell(excel_row, col).font = Font(bold=True)

        if row_type == "region":
            ws.cell(excel_row, 1).value = "   " + str(ws.cell(excel_row, 1).value)

    for row in range(4, ws.max_row + 1):
        for col in [2, 3, 4, 5, 6, 7]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    widths = {
        "A": 28,
        "B": 16,
        "C": 16,
        "D": 13,
        "E": 16,
        "F": 16,
        "G": 13,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"

    wb.save(output_file)
