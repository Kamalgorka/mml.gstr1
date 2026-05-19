import os
import re
import zipfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


WRITE_OFF_OUTPUT_COLUMNS = [
    "ZONE",
    "REGION",
    "BRANCH_STATE",
    "status",
    "cust_id",
    "member_name",
    "mobile_number",
    "od_days",
    "total_arrear",
    "outstanding_principal",
    "outstanding_interest",
]

AMOUNT_COLUMNS = [
    "total_arrear",
    "outstanding_principal",
    "outstanding_interest",
]


def clean_file_name(name):
    name = str(name).replace(".xlsx", "").replace(".xls", "")
    name = re.sub(r'[\\/:*?"<>|]', "", name)
    return name.strip()


def norm_col(col):
    return str(col).strip().lower().replace(" ", "").replace("_", "")


def find_column(columns, required_col):
    target = norm_col(required_col)
    for col in columns:
        if norm_col(col) == target:
            return col
    return None


def normalize_cust_id_series(series):
    return (
        series
        .astype(str)
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )


def get_actual_columns(columns, required_cols, report_name):
    actual_cols = []

    for col in required_cols:
        actual = find_column(columns, col)

        if actual is None:
            raise ValueError(f"Column '{col}' not found in {report_name}")

        actual_cols.append(actual)

    return actual_cols


def apply_red_highlighting(file_path):
    """
    Used only for Write Off Consolidated.xlsx.
    Monthly files are CSV for speed, so no red highlighting there.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    headers = [cell.value for cell in ws[1]]

    if "Discrepancy" not in headers:
        wb.save(file_path)
        return

    discrepancy_col_idx = headers.index("Discrepancy") + 1

    for row in range(2, ws.max_row + 1):
        discrepancy = ws.cell(row=row, column=discrepancy_col_idx).value

        if discrepancy:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

    ws.freeze_panes = "A2"
    wb.save(file_path)


def add_sms_discrepancy_fast(df, check_total_arrear=False):
    principal_col = find_column(df.columns, "outstanding_principal")
    interest_col = find_column(df.columns, "outstanding_interest")
    total_outstanding_col = find_column(df.columns, "total_outstanding")

    if principal_col is None:
        raise ValueError("Column 'outstanding_principal' not found")

    if interest_col is None:
        raise ValueError("Column 'outstanding_interest' not found")

    if total_outstanding_col is None:
        raise ValueError("Column 'total_outstanding' not found")

    required_numeric_cols = [
        principal_col,
        interest_col,
        total_outstanding_col,
    ]

    total_arrear_col = None

    if check_total_arrear:
        total_arrear_col = find_column(df.columns, "total_arrear")

        if total_arrear_col is None:
            total_arrear_col = find_column(df.columns, "total_arrear_sum")

        if total_arrear_col is None:
            raise ValueError("Column 'total_arrear' or 'total_arrear_sum' not found")

        required_numeric_cols.append(total_arrear_col)

    for col in required_numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    calculated_total = df[principal_col] + df[interest_col]

    mismatch_mask = calculated_total.round(2) != df[total_outstanding_col].round(2)
    total_zero_negative_mask = df[total_outstanding_col] <= 0
    principal_negative_mask = df[principal_col] < 0
    interest_negative_mask = df[interest_col] < 0

    discrepancy = pd.Series("", index=df.index, dtype="object")

    discrepancy.loc[mismatch_mask] += (
        "Outstanding principal + interest not matched with total_outstanding; "
    )

    discrepancy.loc[total_zero_negative_mask] += (
        "Total Outstanding is zero/negative; "
    )

    discrepancy.loc[principal_negative_mask] += (
        "outstanding_principal is negative; "
    )

    discrepancy.loc[interest_negative_mask] += (
        "outstanding_interest is negative; "
    )

    if check_total_arrear and total_arrear_col is not None:
        total_arrear_negative_mask = df[total_arrear_col] < 0
        outstanding_less_than_arrear_mask = df[total_outstanding_col] < df[total_arrear_col]

        discrepancy.loc[total_arrear_negative_mask] += (
            f"{total_arrear_col} is negative; "
        )

        discrepancy.loc[outstanding_less_than_arrear_mask] += (
            "total_outstanding is less than total_arrear; "
        )

    df["Discrepancy"] = discrepancy.str.strip("; ")

    return df


def consolidate_jlg_into_il(il_df, jlg_df_list):
    """
    If same cust_id exists in IL and JLG files:
    - Add JLG amount values into IL same cust_id row
    - max_od_days = max of IL/JLG
    - Remove matched cust_id from JLG outputs
    No new consolidated file is created.
    """
    if il_df.empty or not jlg_df_list:
        return il_df, jlg_df_list

    cust_col = find_column(il_df.columns, "cust_id")
    max_od_col = find_column(il_df.columns, "max_od_days")
    total_arrear_col = find_column(il_df.columns, "total_arrear_sum")

    if total_arrear_col is None:
        total_arrear_col = find_column(il_df.columns, "total_arrear")

    principal_col = find_column(il_df.columns, "outstanding_principal")
    interest_col = find_column(il_df.columns, "outstanding_interest")
    total_outstanding_col = find_column(il_df.columns, "total_outstanding")

    required = [
        cust_col,
        max_od_col,
        total_arrear_col,
        principal_col,
        interest_col,
        total_outstanding_col,
    ]

    if any(c is None for c in required):
        raise ValueError("Required consolidation columns missing in IL/JLG data")

    il_df = il_df.copy()
    il_df[cust_col] = normalize_cust_id_series(il_df[cust_col])

    all_jlg = []

    for idx, jlg_df in enumerate(jlg_df_list):
        if jlg_df.empty:
            continue

        temp = jlg_df.copy()
        temp[cust_col] = normalize_cust_id_series(temp[cust_col])
        temp["_source_idx"] = idx
        all_jlg.append(temp)

    if not all_jlg:
        return il_df, jlg_df_list

    jlg_all = pd.concat(all_jlg, ignore_index=True)

    matched_ids = set(il_df[cust_col]).intersection(set(jlg_all[cust_col]))

    if not matched_ids:
        return il_df, jlg_df_list

    amount_cols = [
        max_od_col,
        total_arrear_col,
        principal_col,
        interest_col,
        total_outstanding_col,
    ]

    for col in amount_cols:
        il_df[col] = pd.to_numeric(il_df[col], errors="coerce").fillna(0)
        jlg_all[col] = pd.to_numeric(jlg_all[col], errors="coerce").fillna(0)

    jlg_matched = jlg_all[jlg_all[cust_col].isin(matched_ids)].copy()

    jlg_grouped = (
        jlg_matched
        .groupby(cust_col, as_index=False)
        .agg({
            max_od_col: "max",
            total_arrear_col: "sum",
            principal_col: "sum",
            interest_col: "sum",
            total_outstanding_col: "sum",
        })
    )

    il_df = il_df.merge(
        jlg_grouped,
        on=cust_col,
        how="left",
        suffixes=("", "_jlg")
    )

    for col in [
        total_arrear_col,
        principal_col,
        interest_col,
        total_outstanding_col,
    ]:
        il_df[col] = il_df[col] + il_df[f"{col}_jlg"].fillna(0)
        il_df.drop(columns=[f"{col}_jlg"], inplace=True)

    il_df[max_od_col] = il_df[[max_od_col, f"{max_od_col}_jlg"]].max(axis=1)
    il_df.drop(columns=[f"{max_od_col}_jlg"], inplace=True)

    updated_jlg_list = []

    for jlg_df in jlg_df_list:
        if jlg_df.empty:
            updated_jlg_list.append(jlg_df)
            continue

        updated_df = jlg_df.copy()
        updated_df[cust_col] = normalize_cust_id_series(updated_df[cust_col])
        updated_df = updated_df[~updated_df[cust_col].isin(matched_ids)].copy()
        updated_jlg_list.append(updated_df)

    return il_df, updated_jlg_list


def process_monthly_file(uploaded_file, report_name, od_column_name, output_dir, save_output=True):
    uploaded_file.seek(0)

    header_df = pd.read_excel(uploaded_file, nrows=0)
    all_columns = list(header_df.columns)

    status_col = find_column(all_columns, "status")
    od_col = find_column(all_columns, od_column_name)

    if status_col is None:
        raise ValueError(f"Status column not found in {report_name}")

    if od_col is None:
        raise ValueError(f"{od_column_name} column not found in {report_name}")

    required_check_cols = [
        "outstanding_principal",
        "outstanding_interest",
        "total_outstanding",
    ]

    for col in required_check_cols:
        if find_column(all_columns, col) is None:
            raise ValueError(f"Column '{col}' not found in {report_name}")

    uploaded_file.seek(0)

    df = pd.read_excel(uploaded_file, usecols=all_columns)
    df.columns = [str(c).strip() for c in df.columns]

    status_col = find_column(df.columns, "status")
    od_col = find_column(df.columns, od_column_name)

    cust_col = find_column(df.columns, "cust_id")
    if cust_col is not None:
        df[cust_col] = normalize_cust_id_series(df[cust_col])

    df = df[
        ~df[status_col]
        .astype(str)
        .str.strip()
        .str.lower()
        .str.contains("death", na=False)
    ].copy()

    df[od_col] = pd.to_numeric(df[od_col], errors="coerce").fillna(0)

    arrear_free_df = df[df[od_col] == 0].copy()
    arrear_df = df[df[od_col] > 0].copy()

    arrear_free_df = add_sms_discrepancy_fast(
        arrear_free_df,
        check_total_arrear=False
    )

    arrear_df = add_sms_discrepancy_fast(
        arrear_df,
        check_total_arrear=True
    )

    if save_output:
        safe_name = clean_file_name(report_name)

        arrear_free_path = os.path.join(output_dir, f"Arrear Free {safe_name}.csv")
        arrear_path = os.path.join(output_dir, f"Arrear {safe_name}.csv")

        arrear_free_df.to_csv(arrear_free_path, index=False, encoding="utf-8-sig")
        arrear_df.to_csv(arrear_path, index=False, encoding="utf-8-sig")

    return {
        "report_name": clean_file_name(report_name),
        "type": "monthly",
        "arrear_free_rows": len(arrear_free_df),
        "arrear_rows": len(arrear_df),
        "arrear_free_discrepancy_rows": int((arrear_free_df["Discrepancy"] != "").sum()),
        "arrear_discrepancy_rows": int((arrear_df["Discrepancy"] != "").sum()),
        "total_rows_after_death_removed": len(df),
        "arrear_free_df": arrear_free_df,
        "arrear_df": arrear_df,
    }


def read_writeoff_file(uploaded_file, report_name):
    uploaded_file.seek(0)

    header_df = pd.read_excel(uploaded_file, nrows=0)
    all_columns = list(header_df.columns)

    usecols = get_actual_columns(
        all_columns,
        WRITE_OFF_OUTPUT_COLUMNS,
        report_name
    )

    uploaded_file.seek(0)

    df = pd.read_excel(uploaded_file, usecols=usecols)
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {}

    for required_col in WRITE_OFF_OUTPUT_COLUMNS:
        actual_col = find_column(df.columns, required_col)
        rename_map[actual_col] = required_col

    df = df.rename(columns=rename_map)

    return df[WRITE_OFF_OUTPUT_COLUMNS].copy()


def process_writeoff_files(files, output_dir):
    wo_df = read_writeoff_file(
        files["Loan OS Write Off"],
        "Loan OS Write Off"
    )

    wo_il_df = read_writeoff_file(
        files["Loan OS Write Off IL"],
        "Loan OS Write Off IL"
    )

    df = pd.concat([wo_df, wo_il_df], ignore_index=True)

    df = df[
        df["status"]
        .astype(str)
        .str.strip()
        .str.lower()
        .eq("writeoff")
    ].copy()

    if df.empty:
        output_df = pd.DataFrame(
            columns=WRITE_OFF_OUTPUT_COLUMNS + ["total_outstanding", "Discrepancy"]
        )

    else:
        for col in ["od_days"] + AMOUNT_COLUMNS:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        output_df = (
            df.groupby("cust_id", as_index=False)
            .agg({
                "ZONE": "first",
                "REGION": "first",
                "BRANCH_STATE": "first",
                "status": "first",
                "member_name": "first",
                "mobile_number": "first",
                "od_days": "max",
                "total_arrear": "sum",
                "outstanding_principal": "sum",
                "outstanding_interest": "sum",
            })
        )

        output_df["total_outstanding"] = (
            output_df["outstanding_principal"] +
            output_df["outstanding_interest"]
        )

        def check_discrepancy(row):
            issues = []

            if row["total_outstanding"] <= 0:
                issues.append("Total Outstanding is zero/negative")

            for col in AMOUNT_COLUMNS:
                if row[col] < 0:
                    issues.append(f"{col} is negative")

            if row["total_outstanding"] < row["total_arrear"]:
                issues.append("Total Outstanding is less than Total Arrear")

            return "; ".join(issues)

        output_df["Discrepancy"] = output_df.apply(check_discrepancy, axis=1)

        output_df = output_df[
            [
                "ZONE",
                "REGION",
                "BRANCH_STATE",
                "status",
                "cust_id",
                "member_name",
                "mobile_number",
                "od_days",
                "total_arrear",
                "outstanding_principal",
                "outstanding_interest",
                "total_outstanding",
                "Discrepancy",
            ]
        ]

    output_path = os.path.join(output_dir, "Write Off Consolidated.xlsx")
    output_df.to_excel(output_path, index=False)

    apply_red_highlighting(output_path)

    return {
        "report_name": "Write Off Consolidated",
        "type": "writeoff",
        "total_writeoff_rows_after_filter": len(df),
        "unique_cust_id_rows": len(output_df),
        "discrepancy_rows": int((output_df["Discrepancy"] != "").sum()),
    }


def process_sms_report(files, output_dir, progress_callback=None):
    os.makedirs(output_dir, exist_ok=True)

    monthly_report_config = {
        "Monthly Outstanding SMS Data JLG HUB 1": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 1"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data JLG HUB 2": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 2"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data JLG HUB 3": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 3"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data JLG HUB 4": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 4"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data JLG HUB 5": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 5"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data JLG HUB 6": {
            "file": files["Monthly Outstanding SMS Data JLG HUB 6"],
            "od_col": "max_od_days",
        },
        "Monthly Outstanding SMS Data IL": {
            "file": files["Monthly Outstanding SMS Data IL"],
            "od_col": "max_od_days",
        },
    }

    summary = []
    total_steps = len(monthly_report_config) + 2
    current_step = 0
    monthly_results = {}

    for report_name, cfg in monthly_report_config.items():
        current_step += 1

        if progress_callback:
            progress_callback(
                int(((current_step - 1) / total_steps) * 90),
                f"Processing {current_step}/{total_steps}: {report_name}"
            )

        result = process_monthly_file(
            uploaded_file=cfg["file"],
            report_name=report_name,
            od_column_name=cfg["od_col"],
            output_dir=output_dir,
            save_output=False
        )

        monthly_results[report_name] = result

        if progress_callback:
            progress_callback(
                int((current_step / total_steps) * 90),
                f"Completed {current_step}/{total_steps}: {report_name}"
            )

    jlg_report_names = [
        "Monthly Outstanding SMS Data JLG HUB 1",
        "Monthly Outstanding SMS Data JLG HUB 2",
        "Monthly Outstanding SMS Data JLG HUB 3",
        "Monthly Outstanding SMS Data JLG HUB 4",
        "Monthly Outstanding SMS Data JLG HUB 5",
        "Monthly Outstanding SMS Data JLG HUB 6",
    ]

    il_report_name = "Monthly Outstanding SMS Data IL"

    il_free_df = monthly_results[il_report_name]["arrear_free_df"]
    jlg_free_dfs = [monthly_results[name]["arrear_free_df"] for name in jlg_report_names]

    il_free_df, jlg_free_dfs = consolidate_jlg_into_il(
        il_free_df,
        jlg_free_dfs
    )

    monthly_results[il_report_name]["arrear_free_df"] = il_free_df

    for name, updated_df in zip(jlg_report_names, jlg_free_dfs):
        monthly_results[name]["arrear_free_df"] = updated_df

    il_arrear_df = monthly_results[il_report_name]["arrear_df"]
    jlg_arrear_dfs = [monthly_results[name]["arrear_df"] for name in jlg_report_names]

    il_arrear_df, jlg_arrear_dfs = consolidate_jlg_into_il(
        il_arrear_df,
        jlg_arrear_dfs
    )

    monthly_results[il_report_name]["arrear_df"] = il_arrear_df

    for name, updated_df in zip(jlg_report_names, jlg_arrear_dfs):
        monthly_results[name]["arrear_df"] = updated_df

    for report_name, result in monthly_results.items():
        safe_name = clean_file_name(report_name)

        arrear_free_df = result["arrear_free_df"]
        arrear_df = result["arrear_df"]

        arrear_free_df.to_csv(
            os.path.join(output_dir, f"Arrear Free {safe_name}.csv"),
            index=False,
            encoding="utf-8-sig"
        )

        arrear_df.to_csv(
            os.path.join(output_dir, f"Arrear {safe_name}.csv"),
            index=False,
            encoding="utf-8-sig"
        )

        summary.append({
            "report_name": safe_name,
            "type": "monthly",
            "arrear_free_rows": len(arrear_free_df),
            "arrear_rows": len(arrear_df),
            "arrear_free_discrepancy_rows": int((arrear_free_df["Discrepancy"] != "").sum()),
            "arrear_discrepancy_rows": int((arrear_df["Discrepancy"] != "").sum()),
            "total_rows_after_death_removed": result["total_rows_after_death_removed"],
        })

    current_step += 1

    if progress_callback:
        progress_callback(
            int(((current_step - 1) / total_steps) * 90),
            "Processing Write Off Consolidation"
        )

    writeoff_summary = process_writeoff_files(
        files=files,
        output_dir=output_dir
    )

    summary.append(writeoff_summary)

    if progress_callback:
        progress_callback(95, "Creating ZIP file...")

    zip_path = os.path.join(output_dir, "SMS_Report_Output.zip")

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(output_dir):
            if file.endswith(".csv") or file.endswith(".xlsx"):
                zipf.write(
                    os.path.join(output_dir, file),
                    arcname=file
                )

    if progress_callback:
        progress_callback(100, "Completed.")

    return summary, zip_path
