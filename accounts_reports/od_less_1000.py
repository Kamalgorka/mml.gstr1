import os
import tempfile
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


EXCLUDE_FUNDERS = {
    "ARCILARC_MARCH_2026",
    "CFMARC_MARCH_2025",
    "PHOENIX_ARC",
    "PHOENIX_ARC-1",
}

HUB_MAPPING = {
    "Lucknow_Hub": "Hub-4 Lucknow",
    "Patna_Hub": "Hub-3 Patna",
    "Chandigarh_Hub": "Hub-1 Chandigarh",
    "Bhubaneswar_Hub": "Hub-6 Bhubaneswar",
    "Ahmedabad_Hub": "Hub-2 Ahmedabad",
    "Kolkata_Hub": "Hub-5 Kolkata",
}


def read_uploaded_file(uploaded_file):
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file)

    if filename.endswith(".xlsb"):
        return pd.read_excel(uploaded_file, engine="pyxlsb")

    return pd.read_excel(uploaded_file)


def clean_number(series):
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    ).fillna(0)


def normalize_text(series):
    return series.astype(str).str.strip()


def find_col(df, possible_names):
    normalized = {
        str(c).strip().lower().replace(" ", "_"): c
        for c in df.columns
    }

    for name in possible_names:
        key = name.strip().lower().replace(" ", "_")
        if key in normalized:
            return normalized[key]

    raise ValueError(f"Required column not found. Expected one of: {possible_names}")


def build_hub_region(df, zone_col, state_col):
    df["Hub"] = df[zone_col].map(HUB_MAPPING).fillna(df[zone_col])
    df["Region"] = df[state_col].astype(str).str.strip().str.upper()
    return df


def prepare_filtered_data(arrear_file):
    df = read_uploaded_file(arrear_file)
    df.columns = [str(c).strip() for c in df.columns]

    funder_col = find_col(df, ["funder", "funder_description", "Funder_Description"])
    status_col = find_col(df, ["status"])
    od_days_col = find_col(df, ["od_days", "max_od_days", "OD Days"])
    total_arrear_col = find_col(df, ["total_arrear", "Total Arrear"])
    outstanding_principal_col = find_col(df, ["outstanding_principal", "Outstanding Principal"])
    cust_id_col = find_col(df, ["cust_id", "Cust ID", "Customer ID"])
    zone_col = find_col(df, ["zone", "ZONE", "Hub"])
    state_col = find_col(df, ["state", "STATE", "Region"])

    df[funder_col] = normalize_text(df[funder_col])
    df[status_col] = normalize_text(df[status_col])
    df[od_days_col] = clean_number(df[od_days_col])
    df[total_arrear_col] = clean_number(df[total_arrear_col])
    df[outstanding_principal_col] = clean_number(df[outstanding_principal_col])
    df[cust_id_col] = normalize_text(df[cust_id_col])
    df[zone_col] = normalize_text(df[zone_col])
    df[state_col] = normalize_text(df[state_col])

    # Same first logic as Single Client OD:
    # 1. Exclude selected funders
    # 2. Status Active
    # 3. OD Days > 0
    # 4. New condition: total_arrear <= 1000
    filtered = df[
        ~df[funder_col].isin(EXCLUDE_FUNDERS)
        & df[status_col].str.lower().eq("active")
        & (df[od_days_col] > 0)
        & (df[total_arrear_col] <= 1000)
    ].copy()

    filtered = build_hub_region(filtered, zone_col, state_col)

    return filtered, {
        "total_arrear_col": total_arrear_col,
        "outstanding_principal_col": outstanding_principal_col,
        "cust_id_col": cust_id_col,
    }


def prepare_summary(filtered_df, cols):
    total_arrear_col = cols["total_arrear_col"]
    outstanding_principal_col = cols["outstanding_principal_col"]
    cust_id_col = cols["cust_id_col"]

    summary = (
        filtered_df
        .groupby(["Hub", "Region"], as_index=False)
        .agg(
            Arrear_Amount=(total_arrear_col, "sum"),
            PAR_Amount=(outstanding_principal_col, "sum"),
            No_of_Clients=(cust_id_col, pd.Series.nunique),
        )
    )

    hub_order = [
        "Hub-1 Chandigarh",
        "Hub-2 Ahmedabad",
        "Hub-3 Patna",
        "Hub-4 Lucknow",
        "Hub-5 Kolkata",
        "Hub-6 Bhubaneswar",
    ]

    summary["Hub_Order"] = summary["Hub"].apply(
        lambda x: hub_order.index(x) if x in hub_order else 99
    )

    summary = summary.sort_values(["Hub_Order", "Region"]).drop(columns=["Hub_Order"])

    return summary


def format_number(value):
    try:
        return f"{int(round(value)):,}"
    except Exception:
        return value


def write_summary_sheet(wb, summary_df):
    ws = wb.active
    ws.title = "Summary"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="B7DEE8")
    header_fill = PatternFill("solid", fgColor="B7DEE8")
    hub_fills = {
        "Hub-1 Chandigarh": "DAF2D0",
        "Hub-2 Ahmedabad": "FCE4D6",
        "Hub-3 Patna": "DDEBF7",
        "Hub-4 Lucknow": "E4DFEC",
        "Hub-5 Kolkata": "FFFFCC",
        "Hub-6 Bhubaneswar": "E7E6E6",
    }
    total_fill = PatternFill("solid", fgColor="B7DEE8")

    ws.merge_cells("A1:D1")
    ws["A1"] = "OD Amount <1000 (Non-NPA Clients)"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = title_fill

    headers = ["HUB/ Region", "Arrear Amount", "PAR Amount", "No. of Clients"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 3

    grand_arrear = 0
    grand_par = 0
    grand_clients = 0

    for hub, hub_df in summary_df.groupby("Hub", sort=False):
        fill_color = hub_fills.get(hub, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)

        ws.cell(row=row, column=1, value=hub).font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = border
        row += 1

        hub_arrear = 0
        hub_par = 0
        hub_clients = 0

        for _, r in hub_df.iterrows():
            ws.cell(row=row, column=1, value=r["Region"])
            ws.cell(row=row, column=2, value=format_number(r["Arrear_Amount"]))
            ws.cell(row=row, column=3, value=format_number(r["PAR_Amount"]))
            ws.cell(row=row, column=4, value=format_number(r["No_of_Clients"]))

            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = border

            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")

            hub_arrear += r["Arrear_Amount"]
            hub_par += r["PAR_Amount"]
            hub_clients += r["No_of_Clients"]
            row += 1

        ws.cell(row=row, column=1, value=f"{hub} Total")
        ws.cell(row=row, column=2, value=format_number(hub_arrear))
        ws.cell(row=row, column=3, value=format_number(hub_par))
        ws.cell(row=row, column=4, value=format_number(hub_clients))

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.border = border

        grand_arrear += hub_arrear
        grand_par += hub_par
        grand_clients += hub_clients
        row += 1

    ws.cell(row=row, column=1, value="Grand Total")
    ws.cell(row=row, column=2, value=format_number(grand_arrear))
    ws.cell(row=row, column=3, value=format_number(grand_par))
    ws.cell(row=row, column=4, value=format_number(grand_clients))

    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    for rows in ws.iter_rows():
        for cell in rows:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def process_od_less_1000(arrear_file):
    filtered_df, cols = prepare_filtered_data(arrear_file)
    summary_df = prepare_summary(filtered_df, cols)

    output_dir = tempfile.mkdtemp()
    output_path = os.path.join(output_dir, "OD_Amount_Less_1000_Non_NPA_Clients.xlsx")

    wb = Workbook()
    write_summary_sheet(wb, summary_df)

    detail_ws = wb.create_sheet("Filtered Data")

    for col_idx, col_name in enumerate(filtered_df.columns, start=1):
        detail_ws.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)

    for row_idx, row_data in enumerate(filtered_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            detail_ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(filtered_df.columns) + 1):
        detail_ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(output_path)

    return output_path
