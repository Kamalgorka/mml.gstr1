import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def clean_col(col):
    return str(col).strip().lower().replace(" ", "_")


def read_file(file):
    name = file.name.lower()

    if name.endswith(".csv"):
        return pd.read_csv(file)

    if name.endswith(".xlsb"):
        return pd.read_excel(file, engine="pyxlsb")

    return pd.read_excel(file)


def normalize_branch(x):
    if pd.isna(x):
        return ""
    return str(x).strip().upper().replace("_", "-").replace("  ", " ")


def process_midfin_finpage_recon(
    finpage_jlg_file,
    finpage_il_file,
    midfin_regular_file,
    midfin_od_file,
    branch_master_file,
    output_dir,
    progress_callback=None
):

    if progress_callback:
        progress_callback(10, "Reading uploaded files...")

    # =========================
    # READ FILES
    # =========================
    fin_jlg = read_file(finpage_jlg_file)
    fin_il = read_file(finpage_il_file)
    mid_reg = read_file(midfin_regular_file)
    mid_od = read_file(midfin_od_file)
    branch = read_file(branch_master_file)

    # clean column names
    fin_jlg.columns = [clean_col(c) for c in fin_jlg.columns]
    fin_il.columns = [clean_col(c) for c in fin_il.columns]
    mid_reg.columns = [clean_col(c) for c in mid_reg.columns]
    mid_od.columns = [clean_col(c) for c in mid_od.columns]
    branch.columns = [clean_col(c) for c in branch.columns]

    if progress_callback:
        progress_callback(25, "Preparing Finpage collection data...")

    # =========================
    # FINPAGE DATA
    # JLG: LOAN_ID + AMOUNT_COLLECTED
    # IL : Loan ID + Amount
    # =========================
    required_fin_jlg = ["loan_id", "amount_collected"]
    required_fin_il = ["loan_id", "amount"]

    for col in required_fin_jlg:
        if col not in fin_jlg.columns:
            raise ValueError(f"Finpage JLG file missing column: {col}")

    for col in required_fin_il:
        if col not in fin_il.columns:
            raise ValueError(f"Finpage IL file missing column: {col}")

    fin_jlg_data = fin_jlg[["loan_id", "amount_collected"]].copy()
    fin_jlg_data.rename(columns={"amount_collected": "amount_received_in_finpage"}, inplace=True)

    fin_il_data = fin_il[["loan_id", "amount"]].copy()
    fin_il_data.rename(columns={"amount": "amount_received_in_finpage"}, inplace=True)

    finpage = pd.concat([fin_jlg_data, fin_il_data], ignore_index=True)

    finpage["loan_id"] = finpage["loan_id"].astype(str).str.strip()
    finpage["amount_received_in_finpage"] = pd.to_numeric(
        finpage["amount_received_in_finpage"], errors="coerce"
    ).fillna(0)

    finpage_final = (
        finpage.groupby("loan_id", as_index=False)["amount_received_in_finpage"]
        .sum()
    )

    if progress_callback:
        progress_callback(45, "Preparing Midfin collection data...")

    # =========================
    # MIDFIN DATA
    # Both reports same heading:
    # LOAN NUMBER + TOTAL AMOUNT COLLECTED + LOAN BRANCH
    # =========================
    required_mid_cols = ["loan_number", "total_amount_collected", "loan_branch"]

    for df_name, df in [("Midfin Regular Collection", mid_reg), ("Midfin OD Collection", mid_od)]:
        for col in required_mid_cols:
            if col not in df.columns:
                raise ValueError(f"{df_name} file missing column: {col}")

    mid_reg_data = mid_reg[["loan_number", "total_amount_collected", "loan_branch"]].copy()
    mid_od_data = mid_od[["loan_number", "total_amount_collected", "loan_branch"]].copy()

    midfin = pd.concat([mid_reg_data, mid_od_data], ignore_index=True)

    midfin["loan_number"] = midfin["loan_number"].astype(str).str.strip()
    midfin["total_amount_collected"] = pd.to_numeric(
        midfin["total_amount_collected"], errors="coerce"
    ).fillna(0)

    # Keep branch against loan number
    branch_ref = (
        midfin[["loan_number", "loan_branch"]]
        .dropna(subset=["loan_number"])
        .drop_duplicates(subset=["loan_number"], keep="first")
    )

    midfin_final = (
        midfin.groupby("loan_number", as_index=False)["total_amount_collected"]
        .sum()
    )

    midfin_final = midfin_final.merge(branch_ref, on="loan_number", how="left")

    if progress_callback:
        progress_callback(65, "Merging Branch Master...")

    # =========================
    # BRANCH MASTER
    # Region | Branch Code | Branch Name | Maker | Checker
    # =========================
    required_branch_cols = ["region", "branch_code", "branch_name", "maker", "checker"]

    for col in required_branch_cols:
        if col not in branch.columns:
            raise ValueError(f"Branch Master missing column: {col}")

    branch_master = branch[required_branch_cols].copy()
    branch_master["branch_key"] = branch_master["branch_name"].apply(normalize_branch)

    # =========================
    # SUMMARY PREPARATION
    # =========================
    summary = midfin_final.copy()

    summary.rename(columns={
        "loan_number": "Loan ID",
        "total_amount_collected": "Amount Received in Midfincollect",
        "loan_branch": "Branch Details"
    }, inplace=True)

    summary["loan_id_key"] = summary["Loan ID"].astype(str).str.strip()

    summary = summary.merge(
        finpage_final,
        left_on="loan_id_key",
        right_on="loan_id",
        how="left"
    )

    summary["amount_received_in_finpage"] = summary["amount_received_in_finpage"].fillna(0)

    summary.rename(columns={
        "amount_received_in_finpage": "Amount Received in Finpage"
    }, inplace=True)

    summary["Difference"] = (
        summary["Amount Received in Midfincollect"]
        - summary["Amount Received in Finpage"]
    )

    summary["branch_key"] = summary["Branch Details"].apply(normalize_branch)

    summary = summary.merge(
        branch_master,
        on="branch_key",
        how="left"
    )

    summary["Concerned Branch Executive Remarks"] = summary["Difference"].apply(
        lambda x: "Reconciled" if round(float(x), 2) == 0 else ""
    )

    # Final column order
    summary = summary[[
        "Loan ID",
        "branch_code",
        "Branch Details",
        "maker",
        "checker",
        "Amount Received in Midfincollect",
        "Amount Received in Finpage",
        "Difference",
        "Concerned Branch Executive Remarks"
    ]]

    summary.rename(columns={
        "branch_code": "B.Code",
        "maker": "Branch Accounts Executive",
        "checker": "Concerned TL Name"
    }, inplace=True)

    if progress_callback:
        progress_callback(85, "Creating Excel output...")

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "Midfin_and_Finpage_Collection_Recon.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        midfin_final.to_excel(writer, sheet_name="Midfin Consolidated", index=False)
        finpage_final.to_excel(writer, sheet_name="Finpage Consolidated", index=False)

    # =========================
    # FORMATTING
    # =========================
    wb = load_workbook(output_path)

    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="000000")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

        ws.freeze_panes = "A2"

    wb.save(output_path)

    if progress_callback:
        progress_callback(100, "Report generated successfully.")

    return output_path
