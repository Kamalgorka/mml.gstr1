import os
import tempfile
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


EXCLUDE_FUNDERS = {
    "ARCILARC_MARCH_2026",
    "CFMARC_MARCH_2025",
    "PHOENIX_ARC",
    "PHOENIX_ARC-1",
}

HUB_MAPPING = {
    "Lucknow_Hub": "Hub-4 Lucknow",
    "Patna_Hub": "Hub-3 Patna",
    "Chandigarh_Hub": "Hub-1 Chandigarh",
    "Bhubaneswar_Hub": "Hub-6 Bhubaneswar",
    "Ahmedabad_Hub": "Hub-2 Ahmedabad",
    "Kolkata_Hub": "Hub-5 Kolkata",
}


def read_uploaded_file(uploaded_file):
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file)

    if filename.endswith(".xlsb"):
        return pd.read_excel(uploaded_file, engine="pyxlsb")

    return pd.read_excel(uploaded_file)


def process_single_client_od(uploaded_file):
    df = read_uploaded_file(uploaded_file)

    required_cols = [
        "status",
        "funder",
        "od_days",
        "center_name",
        "cust_id",
        "zone_name",
        "cluster_name",
        "total_arrear",
        "outstanding_principal",
    ]

    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise KeyError(f"Missing required columns: {missing_cols}")

    df["status"] = df["status"].astype(str).str.strip()
    df["funder"] = df["funder"].astype(str).str.strip()
    df["center_name"] = df["center_name"].astype(str).str.strip()
    df["cust_id"] = df["cust_id"].astype(str).str.strip()
    df["zone_name"] = df["zone_name"].astype(str).str.strip()
    df["cluster_name"] = df["cluster_name"].astype(str).str.strip()

    df["od_days"] = pd.to_numeric(df["od_days"], errors="coerce")
    df["total_arrear"] = pd.to_numeric(df["total_arrear"], errors="coerce").fillna(0)
    df["outstanding_principal"] = pd.to_numeric(df["outstanding_principal"], errors="coerce").fillna(0)

    df = df[~df["status"].str.upper().eq("DEATH")]
    df = df[~df["funder"].str.upper().isin({x.upper() for x in EXCLUDE_FUNDERS})]
    df = df[df["od_days"] <= 90]

    df = df[
        (df["center_name"] != "")
        & (df["cust_id"] != "")
        & (df["center_name"].str.lower() != "nan")
        & (df["cust_id"].str.lower() != "nan")
    ]

    center_counts = df.groupby("center_name")["cust_id"].nunique()
    single_centers = center_counts[center_counts == 1].index

    final_df = df[df["center_name"].isin(single_centers)].copy()

    final_df["Hub_Name"] = (
        final_df["zone_name"]
        .map(HUB_MAPPING)
        .fillna(final_df["zone_name"])
    )

    cluster_summary = (
        final_df.groupby(["Hub_Name", "cluster_name"], dropna=False)
        .agg(
            Arrear_Amount=("total_arrear", "sum"),
            PAR_Amount=("outstanding_principal", "sum"),
            No_of_Centers=("center_name", pd.Series.nunique),
            No_of_Clients=("cust_id", pd.Series.nunique),
        )
        .reset_index()
    )

    hub_totals = (
        final_df.groupby("Hub_Name", dropna=False)
        .agg(
            Arrear_Amount=("total_arrear", "sum"),
            PAR_Amount=("outstanding_principal", "sum"),
            No_of_Centers=("center_name", pd.Series.nunique),
            No_of_Clients=("cust_id", pd.Series.nunique),
        )
        .reset_index()
    )

    hub_order = [
        "Hub-1 Chandigarh",
        "Hub-2 Ahmedabad",
        "Hub-3 Patna",
        "Hub-4 Lucknow",
        "Hub-5 Kolkata",
        "Hub-6 Bhubaneswar",
    ]

    summary_rows = [
        ["Single Client OD in a Center (Non-NPA Clients)", "", "", "", ""],
        ["HUB/ Region", "Arrear Amount", "PAR Amount", "No. of Centers", "No. of Clients"],
    ]

    for hub in hub_order:
        hub_data = cluster_summary[cluster_summary["Hub_Name"] == hub].copy()

        if hub_data.empty:
            continue

        hub_data = hub_data.sort_values("cluster_name")

        summary_rows.append([hub, "", "", "", ""])

        for _, row in hub_data.iterrows():
            summary_rows.append([
                row["cluster_name"],
                row["Arrear_Amount"],
                row["PAR_Amount"],
                row["No_of_Centers"],
                row["No_of_Clients"],
            ])

        total_row = hub_totals[hub_totals["Hub_Name"] == hub].iloc[0]

        summary_rows.append([
            f"{hub} Total",
            total_row["Arrear_Amount"],
            total_row["PAR_Amount"],
            total_row["No_of_Centers"],
            total_row["No_of_Clients"],
        ])

    summary_rows.append([
        "Grand Total",
        final_df["total_arrear"].sum(),
        final_df["outstanding_principal"].sum(),
        final_df["center_name"].nunique(),
        final_df["cust_id"].nunique(),
    ])

    summary_df = pd.DataFrame(summary_rows)

    output_dir = tempfile.mkdtemp()
    output_path = os.path.join(output_dir, "Single Client OD in a Center.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False, header=False)
        final_df.to_excel(writer, sheet_name="Member wise data", index=False)

    wb = load_workbook(output_path)
    ws = wb["Summary"]

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    title_fill = PatternFill("solid", fgColor="B7DEE8")
    header_fill = PatternFill("solid", fgColor="B7DEE8")
    hub_fills = [
        "D9EAD3",
        "FCE4D6",
        "DDEBF7",
        "E4DFEC",
        "FFFFCC",
        "EDEDED",
    ]
    grand_fill = PatternFill("solid", fgColor="B7DEE8")

    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = title_fill

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    current_hub_fill = PatternFill("solid", fgColor="D9EAD3")
    hub_index = -1

    for row in range(1, ws.max_row + 1):
        first_value = str(ws.cell(row=row, column=1).value or "")

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col in [2, 3, 4, 5] and row > 2:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        if first_value.startswith("Hub-") and "Total" not in first_value:
            hub_index += 1
            current_hub_fill = PatternFill(
                "solid",
                fgColor=hub_fills[hub_index % len(hub_fills)]
            )
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = current_hub_fill
                ws.cell(row=row, column=col).font = Font(bold=True)

        elif first_value.startswith("Hub-") and "Total" in first_value:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = current_hub_fill
                ws.cell(row=row, column=col).font = Font(bold=True)

        elif first_value == "Grand Total":
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = grand_fill
                ws.cell(row=row, column=col).font = Font(bold=True)

        elif row > 2:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = current_hub_fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    wb.active = wb.sheetnames.index("Summary")
    wb.save(output_path)

    return output_path
